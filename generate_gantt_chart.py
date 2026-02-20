"""
手術室ガントチャート生成スクリプト
===================================
手術実施データ（Excel）を読み込み、日付ごとに手術室の稼働状況を
ガントチャートとしてExcelファイルに出力します。

使い方:
    python generate_gantt_chart.py

入力: ガントチャート-元データ.xlsx（同一フォルダに配置）
出力: 手術室ガントチャート-結果.xlsx（同一フォルダに生成）
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime, timedelta
import os
import sys

# ========== 設定 ==========
# PyInstaller exe の場合は exe の場所、通常実行の場合はスクリプトの場所を基準にする
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_FILE = os.path.join(BASE_DIR, "ガントチャート-元データ.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "手術室ガントチャート-結果.xlsx")

# 手術室の表示順
ROOM_ORDER = ["01A", "01B", "02", "03", "05", "06", "07", "08", "09", "10", "ｱﾝｷﾞｵ"]

# 時間範囲: 8:00 ~ 22:00（10分刻み）
TIME_START_HOUR = 8
TIME_END_HOUR = 22
COLS_PER_HOUR = 6  # 1時間=6列（10分刻み）

# テンプレート行オフセット（テンプレートの6行目=ヘッダ、7~17行目=部屋行）
TPL_HEADER_ROW = 6
TPL_FIRST_ROOM_ROW = 7
TPL_LAST_ROOM_ROW = 17
TPL_COL_START = 2   # B列
TPL_COL_END = 93    # CO列

# 診療科の略称マッピング
DEPT_SHORT = {
    "総合外科": "総",
    "乳腺外科": "乳",
    "心臓血管外科": "心",
    "整形外科・スポーツ診療科": "整",
    "形成外科": "形",
    "産科・婦人科": "産",
    "腎・高血圧内科": "腎",
    "小児外科": "小",
    "眼科": "眼",
    "循環器内科": "循",
    "呼吸器外科": "呼",
    "泌尿器科": "泌",
    "耳鼻咽喉・頭頚科": "耳",
    "脳神経外科": "脳",
    "麻酔科・ペインクリニック": "麻",
}

# 実施区分の色分け（デフォルト値。テンプレートシートがあれば上書きされる）
COLOR_SCHEDULED = "A0C8E4"   # 水色（定時）  ← テンプレートC2
COLOR_URGENT = "6DABD5"      # 濃い水色（臨時） ← テンプレートC3
COLOR_EMERGENCY = "FF8CCC"   # ピンク（緊急）  ← テンプレートC4

# フォント設定
FONT_NAME = "Meiryo UI"

# ラベルフォント（デフォルト値。テンプレートシートC5があれば上書きされる）
LABEL_FONT_NAME = FONT_NAME
LABEL_FONT_SIZE = 6

# テンプレートから読み取った書式情報を格納するグローバル変数
TPL_COL_WIDTHS = {}      # {col_letter: width}
TPL_ROW_HEIGHTS = {}     # {tpl_row_offset: height}  (0=ヘッダ, 1~11=部屋行)
TPL_BORDERS = {}         # {(tpl_row_offset, col): Border}
TPL_HEADER_CELLS = {}    # {col: {font, alignment, value}}
TPL_DATE_FONT = None     # B7セルのフォント
TPL_DATE_ALIGNMENT = None
TPL_ROOM_FONT = None     # C7セルのフォント
TPL_HAS_TEMPLATE = False


def load_template(tpl_ws):
    """テンプレートシートB6:CO17から書式情報を読み取る"""
    global TPL_COL_WIDTHS, TPL_ROW_HEIGHTS, TPL_BORDERS
    global TPL_HEADER_CELLS, TPL_DATE_FONT, TPL_DATE_ALIGNMENT, TPL_ROOM_FONT
    global TPL_HAS_TEMPLATE

    TPL_HAS_TEMPLATE = True

    # 列幅（B～CO）
    # openpyxlは範囲指定の列幅(min~max)を先頭列にしか反映しないため、
    # worksheet.column_dimensions内部データから範囲指定を正しく展開する
    for key, dim in tpl_ws.column_dimensions.items():
        if dim.width is None:
            continue
        if hasattr(dim, 'min') and hasattr(dim, 'max') and dim.min and dim.max:
            for col_idx in range(dim.min, dim.max + 1):
                if TPL_COL_START <= col_idx <= TPL_COL_END:
                    TPL_COL_WIDTHS[get_column_letter(col_idx)] = dim.width
        else:
            col_idx = dim.min if hasattr(dim, 'min') and dim.min else None
            if col_idx and TPL_COL_START <= col_idx <= TPL_COL_END:
                TPL_COL_WIDTHS[get_column_letter(col_idx)] = dim.width

    # 行高（6～17 → offset 0～11）
    for r in range(TPL_HEADER_ROW, TPL_LAST_ROOM_ROW + 1):
        offset = r - TPL_HEADER_ROW
        h = tpl_ws.row_dimensions[r].height
        if h:
            TPL_ROW_HEIGHTS[offset] = h

    # 罫線（B6:CO17 → offset, col で格納）
    for r in range(TPL_HEADER_ROW, TPL_LAST_ROOM_ROW + 1):
        offset = r - TPL_HEADER_ROW
        for c in range(TPL_COL_START, TPL_COL_END + 1):
            cell = tpl_ws.cell(row=r, column=c)
            TPL_BORDERS[(offset, c)] = copy(cell.border)

    # 6行目ヘッダセル（フォント・配置・値）
    for c in range(TPL_COL_START, TPL_COL_END + 1):
        cell = tpl_ws.cell(row=TPL_HEADER_ROW, column=c)
        TPL_HEADER_CELLS[c] = {
            'font': copy(cell.font),
            'alignment': copy(cell.alignment),
            'value': cell.value,
        }

    # B7セル（日付フォント・配置）
    b7 = tpl_ws.cell(row=TPL_FIRST_ROOM_ROW, column=2)
    TPL_DATE_FONT = copy(b7.font)
    TPL_DATE_ALIGNMENT = copy(b7.alignment)

    # C7セル（部屋名フォント）
    c7 = tpl_ws.cell(row=TPL_FIRST_ROOM_ROW, column=3)
    TPL_ROOM_FONT = copy(c7.font)

    print("テンプレートB6:CO17から書式情報を読み取りました")


def time_to_col(time_val, col_offset=4):
    """時刻をExcel列番号に変換（列Dが8:00開始）"""
    import datetime as dt_module
    if isinstance(time_val, str):
        parts = time_val.split(":")
        hours, minutes = int(parts[0]), int(parts[1])
    elif isinstance(time_val, timedelta):
        total_seconds = int(time_val.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
    elif isinstance(time_val, dt_module.time):
        hours = time_val.hour
        minutes = time_val.minute
    else:
        hours = time_val.hour
        minutes = time_val.minute

    total_minutes = (hours - TIME_START_HOUR) * 60 + minutes
    col = col_offset + int(total_minutes / 10)
    return col


def shorten_surgery_name(name, max_chars=20):
    """手術名を短縮"""
    if not name or not isinstance(name, str):
        return ""
    result = name
    if len(result) > max_chars:
        result = result[:max_chars]
    return result


def calculate_utilization(day_data, rooms, weekday=""):
    """稼働率を計算"""
    import datetime as dt_module

    if "土" in weekday:
        calc_start = 9 * 60
        calc_end = 13 * 60
        standard_minutes = 240
    else:
        calc_start = 9 * 60
        calc_end = 17 * 60
        standard_minutes = 480

    room_count = 9.0
    total_available = standard_minutes * room_count
    total_used = 0

    ROOM_WEIGHT = {
        "01A": 0.5,
        "01B": 0.5,
        "ｱﾝｷﾞｵ": 0,
    }

    for _, row in day_data.iterrows():
        try:
            room_name = str(row["実施手術室名"])
            weight = ROOM_WEIGHT.get(room_name, 1.0)
            if weight == 0:
                continue

            t = row["入室時刻"]
            if isinstance(t, str):
                parts = t.split(":")
                start_min = int(parts[0]) * 60 + int(parts[1])
            elif isinstance(t, timedelta):
                start_min = int(t.total_seconds()) // 60
            elif isinstance(t, dt_module.time):
                start_min = t.hour * 60 + t.minute
            else:
                start_min = t.hour * 60 + t.minute

            t2 = row["麻酔終了時刻"]
            if isinstance(t2, str):
                parts = t2.split(":")
                end_min = int(parts[0]) * 60 + int(parts[1])
            elif isinstance(t2, timedelta):
                end_min = int(t2.total_seconds()) // 60
            elif isinstance(t2, dt_module.time):
                end_min = t2.hour * 60 + t2.minute
            else:
                end_min = t2.hour * 60 + t2.minute

            clipped_start = max(start_min, calc_start)
            clipped_end = min(end_min, calc_end)
            if clipped_end > clipped_start:
                total_used += (clipped_end - clipped_start) * weight
        except Exception:
            pass

    if total_available > 0:
        return total_used / total_available
    return 0


def get_tpl_border(row_offset, col):
    """テンプレートの罫線を取得（なければ空Border）"""
    if TPL_HAS_TEMPLATE and (row_offset, col) in TPL_BORDERS:
        return copy(TPL_BORDERS[(row_offset, col)])
    return Border()


def merge_border_with_fill(tpl_border):
    """テンプレート罫線をコピーして返す（塗りつぶし時に罫線を保持するため）"""
    return copy(tpl_border)


def write_day_block(ws, start_row, date_str, weekday, day_data, rooms):
    """1日分のガントチャートブロックを書き込む"""

    header_row = start_row
    utilization = calculate_utilization(day_data, rooms, weekday)
    last_col = TPL_COL_END  # CO列=93

    # 行高を設定
    if TPL_HAS_TEMPLATE:
        for offset, h in TPL_ROW_HEIGHTS.items():
            ws.row_dimensions[start_row + offset].height = h

    # --- ヘッダ行（時間軸） ---
    # テンプレートの6行目の書式を適用
    for c in range(TPL_COL_START, TPL_COL_END + 1):
        cell = ws.cell(row=header_row, column=c)
        if TPL_HAS_TEMPLATE and c in TPL_HEADER_CELLS:
            hdr = TPL_HEADER_CELLS[c]
            cell.font = copy(hdr['font'])
            cell.alignment = copy(hdr['alignment'])
        cell.border = get_tpl_border(0, c)

    # B6: "日付"
    ws.cell(row=header_row, column=2, value="日付")

    # C6: "部屋名"
    ws.cell(row=header_row, column=3, value="部屋名")

    # 時間ヘッダ（8:00～22:00、各6列結合）
    for h in range(TIME_START_HOUR, TIME_END_HOUR + 1):
        col = 4 + (h - TIME_START_HOUR) * COLS_PER_HOUR
        time_label = h * 100
        ws.cell(row=header_row, column=col, value=time_label)
        end_col = col + COLS_PER_HOUR - 1
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=end_col)

    # --- 部屋ごとの行 ---
    for room_idx, room in enumerate(rooms):
        row = start_row + 1 + room_idx
        room_data = day_data[day_data["実施手術室名"] == room]
        tpl_row_offset = 1 + room_idx  # テンプレートの7行目~17行目に対応

        # 罫線をテンプレートから適用
        for c in range(TPL_COL_START, TPL_COL_END + 1):
            ws.cell(row=row, column=c).border = get_tpl_border(tpl_row_offset, c)

        # 日付列（最初の部屋行のみ表示、全部屋を縦結合）
        if room_idx == 0:
            util_label = f"{date_str}\n{utilization:.1%}"
            date_cell = ws.cell(row=row, column=2, value=util_label)
            if TPL_HAS_TEMPLATE and TPL_DATE_FONT:
                date_cell.font = copy(TPL_DATE_FONT)
                date_cell.alignment = copy(TPL_DATE_ALIGNMENT)
            else:
                date_cell.font = Font(name=FONT_NAME, size=9, bold=True)
                date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            date_cell.border = get_tpl_border(tpl_row_offset, 2)
            if len(rooms) > 1:
                ws.merge_cells(start_row=row, start_column=2, end_row=row + len(rooms) - 1, end_column=2)

        # 部屋名
        room_cell = ws.cell(row=row, column=3, value=room)
        if TPL_HAS_TEMPLATE and TPL_ROOM_FONT:
            room_cell.font = copy(TPL_ROOM_FONT)
        else:
            room_cell.font = Font(name=FONT_NAME, size=7)
        room_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 手術バーを描画
        for _, op in room_data.iterrows():
            try:
                start_col = time_to_col(op["入室時刻"])
                end_col = time_to_col(op["麻酔終了時刻"])

                start_col = max(start_col, 4)
                end_col = min(end_col, last_col)

                if end_col <= start_col:
                    end_col = start_col + 1

                dept_short = DEPT_SHORT.get(op["執刀診療科名"], op["執刀診療科名"][0])

                urgency = op.get("実施申込区分", "定時")
                if urgency == "緊急":
                    color = COLOR_EMERGENCY
                elif urgency == "臨時":
                    color = COLOR_URGENT
                else:
                    color = COLOR_SCHEDULED
                fill = PatternFill('solid', fgColor=color)

                surgery_name = op.get("実施手術名０１", "")
                if not isinstance(surgery_name, str):
                    surgery_name = ""
                short_name = shorten_surgery_name(surgery_name, max_chars=40)

                bar_label = f"【{dept_short}】-{short_name}"
                bar_font = Font(name=LABEL_FONT_NAME, size=LABEL_FONT_SIZE, color="000000")

                for c in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = fill
                    # 塗りつぶし後もテンプレート罫線を保持
                    cell.border = get_tpl_border(tpl_row_offset, c)

                ws.cell(row=row, column=start_col, value=bar_label)
                ws.cell(row=row, column=start_col).font = bar_font
                ws.cell(row=row, column=start_col).alignment = Alignment(vertical='center', wrap_text=False)

            except Exception:
                pass

    return start_row + 1 + len(rooms)


def setup_gantt_sheet(ws, title):
    """ガントチャートシートの共通初期設定（列幅・タイトル・凡例）"""
    # 列幅をテンプレートから適用
    ws.column_dimensions['A'].width = 2
    if TPL_HAS_TEMPLATE:
        for letter, w in TPL_COL_WIDTHS.items():
            ws.column_dimensions[letter].width = w
    else:
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 6
        total_time_cols = (TIME_END_HOUR - TIME_START_HOUR + 1) * COLS_PER_HOUR
        for i in range(4, 4 + total_time_cols):
            ws.column_dimensions[get_column_letter(i)].width = 2.5

    ws.cell(row=1, column=2, value=title)
    ws.cell(row=1, column=2).font = Font(name=FONT_NAME, size=14, bold=True)

    legend_row = 2
    legend_col = 2
    ws.cell(row=legend_row, column=legend_col, value="■凡例:")
    ws.cell(row=legend_row, column=legend_col).font = Font(name=FONT_NAME, size=8, bold=True)

    cell = ws.cell(row=legend_row, column=legend_col + 2, value="定時")
    cell.fill = PatternFill('solid', fgColor=COLOR_SCHEDULED)
    cell.font = Font(name=FONT_NAME, size=8)
    cell.alignment = Alignment(horizontal='center')

    cell = ws.cell(row=legend_row, column=legend_col + 4, value="臨時")
    cell.fill = PatternFill('solid', fgColor=COLOR_URGENT)
    cell.font = Font(name=FONT_NAME, size=8)
    cell.alignment = Alignment(horizontal='center')

    cell = ws.cell(row=legend_row, column=legend_col + 6, value="緊急")
    cell.fill = PatternFill('solid', fgColor=COLOR_EMERGENCY)
    cell.font = Font(name=FONT_NAME, size=8)
    cell.alignment = Alignment(horizontal='center')

    ws.cell(row=legend_row + 1, column=legend_col, value="※稼働率 = 平日:9:00-17:00（8h×9室）、土曜:9:00-13:00（4h×9室）")
    ws.cell(row=legend_row + 1, column=legend_col).font = Font(name=FONT_NAME, size=8)
    ws.cell(row=legend_row + 2, column=legend_col, value="※01A・01Bは各0.5室換算、アンギオ室は除外")
    ws.cell(row=legend_row + 2, column=legend_col).font = Font(name=FONT_NAME, size=8)

    ws.sheet_view.zoomScale = 80
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def write_gantt_for_dates(ws, df, date_list, weekday_map):
    """日付リストに従ってガントチャートブロックを書き込む"""
    current_row = 6
    count = 0
    for date_str in date_list:
        day_data = df[df["手術実施日"] == date_str]
        weekday = weekday_map.get(date_str, "")
        weekday_short = weekday.replace("曜日", "") if isinstance(weekday, str) else ""

        try:
            dt = pd.to_datetime(date_str)
            date_display = f"{dt.month:02d}/{dt.day:02d}({weekday_short})"
        except Exception:
            date_display = date_str

        next_row = write_day_block(ws, current_row, date_display, weekday_short, day_data, ROOM_ORDER)
        current_row = next_row + 1
        count += 1
    return count


def main():
    print(f"入力ファイル読み込み: {INPUT_FILE}")
    df = pd.read_excel(INPUT_FILE, sheet_name="ガントチャートデータ", dtype={"実施手術室名": str})

    # 日付でソート
    df["手術実施日_sort"] = pd.to_datetime(df["手術実施日"], format="%Y/%m/%d")
    df = df.sort_values(["手術実施日_sort", "実施手術室名", "入室時刻"])

    dates = df["手術実施日"].unique()
    weekday_map = dict(zip(df["手術実施日"], df["曜日"]))

    wb = Workbook()

    # === シート1: ガントチャートデータ（元データコピー） ===
    data_ws = wb.active
    data_ws.title = "ガントチャートデータ"

    # テンプレートシートから設定を読み取り
    global COLOR_SCHEDULED, COLOR_URGENT, COLOR_EMERGENCY, LABEL_FONT_NAME, LABEL_FONT_SIZE
    src_wb = load_workbook(INPUT_FILE)
    if "テンプレート" in src_wb.sheetnames:
        tpl_ws = src_wb["テンプレート"]

        # 色の読み取り（C2=定時、C3=臨時、C4=緊急）
        c2_fill = tpl_ws.cell(row=2, column=3).fill
        if c2_fill.fill_type == "solid" and c2_fill.fgColor and c2_fill.fgColor.rgb:
            rgb = str(c2_fill.fgColor.rgb)
            if len(rgb) == 8:
                rgb = rgb[2:]
            COLOR_SCHEDULED = rgb
            print(f"テンプレートC2から定時の色を取得: #{COLOR_SCHEDULED}")
        c3_fill = tpl_ws.cell(row=3, column=3).fill
        if c3_fill.fill_type == "solid" and c3_fill.fgColor and c3_fill.fgColor.rgb:
            rgb = str(c3_fill.fgColor.rgb)
            if len(rgb) == 8:
                rgb = rgb[2:]
            COLOR_URGENT = rgb
            print(f"テンプレートC3から臨時の色を取得: #{COLOR_URGENT}")
        c4_fill = tpl_ws.cell(row=4, column=3).fill
        if c4_fill.fill_type == "solid" and c4_fill.fgColor and c4_fill.fgColor.rgb:
            rgb = str(c4_fill.fgColor.rgb)
            if len(rgb) == 8:
                rgb = rgb[2:]
            COLOR_EMERGENCY = rgb
            print(f"テンプレートC4から緊急の色を取得: #{COLOR_EMERGENCY}")

        # ラベルフォントの読み取り
        c5_font = tpl_ws.cell(row=5, column=3).font
        if c5_font.name:
            LABEL_FONT_NAME = c5_font.name
        if c5_font.size:
            LABEL_FONT_SIZE = c5_font.size
        print(f"テンプレートC5からラベルフォントを取得: {LABEL_FONT_NAME}, {LABEL_FONT_SIZE}pt")

        # 書式情報の読み取り（列幅・行高・罫線・フォント）
        load_template(tpl_ws)

    # ガントチャートデータシートのコピー
    if "ガントチャートデータ" in src_wb.sheetnames:
        src_ws = src_wb["ガントチャートデータ"]

        for col_letter, dim in src_ws.column_dimensions.items():
            data_ws.column_dimensions[col_letter].width = dim.width
            data_ws.column_dimensions[col_letter].hidden = dim.hidden

        for row_num, dim in src_ws.row_dimensions.items():
            data_ws.row_dimensions[row_num].height = dim.height
            data_ws.row_dimensions[row_num].hidden = dim.hidden

        for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
            for cell in row:
                dst_cell = data_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    dst_cell.font = copy(cell.font)
                    dst_cell.fill = copy(cell.fill)
                    dst_cell.border = copy(cell.border)
                    dst_cell.alignment = copy(cell.alignment)
                    dst_cell.number_format = cell.number_format

        for merged_range in src_ws.merged_cells.ranges:
            data_ws.merge_cells(str(merged_range))

    src_wb.close()

    # === シート2: 手術室ガントチャート（日付順） ===
    ws_date = wb.create_sheet("手術室ガントチャート")
    setup_gantt_sheet(ws_date, "手術室 ガントチャート（2025年9月）")
    count_date = write_gantt_for_dates(ws_date, df, dates, weekday_map)

    # === シート3: 手術室ガントチャート・曜日順 ===
    WEEKDAY_ORDER = {"月": 0, "火": 1, "水": 2, "木": 3, "金": 4, "土": 5, "日": 6}

    date_info = []
    for date_str in dates:
        weekday = weekday_map.get(date_str, "")
        weekday_short = weekday.replace("曜日", "") if isinstance(weekday, str) else ""
        try:
            dt = pd.to_datetime(date_str)
            nth = (dt.day - 1) // 7 + 1
        except Exception:
            nth = 1
        wday_order = WEEKDAY_ORDER.get(weekday_short, 9)
        date_info.append((wday_order, nth, date_str))

    date_info.sort(key=lambda x: (x[0], x[1]))
    dates_by_weekday = [d[2] for d in date_info]

    ws_weekday = wb.create_sheet("手術室ガントチャート・曜日順")
    setup_gantt_sheet(ws_weekday, "手術室 ガントチャート・曜日順（2025年9月）")
    write_gantt_for_dates(ws_weekday, df, dates_by_weekday, weekday_map)

    # 保存
    wb.save(OUTPUT_FILE)
    print(f"ガントチャート生成完了: {OUTPUT_FILE}")
    print(f"全{count_date}日分のガントチャートを出力しました。")


if __name__ == "__main__":
    main()
